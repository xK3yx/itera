from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete, func
from uuid import UUID
import io

try:
    from pypdf import PdfReader as _PdfReader
except ImportError:
    _PdfReader = None

try:
    from docx import Document as _DocxDocument
except ImportError:
    _DocxDocument = None

try:
    from openpyxl import load_workbook as _load_workbook
except ImportError:
    _load_workbook = None

from app.database import get_db
from app.models.user import User
from app.models.session import Session
from app.models.message import Message
from app.models.roadmap import Roadmap
from app.middleware.auth_middleware import get_current_user
from app.services.ai_service import ai_service
from app.services.session_service import (
    get_conversation_history,
    save_message,
    save_roadmap
)
from app.schemas.chat import (
    StartSessionRequest,
    StartSessionResponse,
    SendMessageRequest,
    SendMessageResponse,
    SessionHistoryResponse,
    MessageResponse
)

router = APIRouter(prefix="/api/v1/chat", tags=["Chat"])


@router.post("/start", response_model=StartSessionResponse)
async def start_session(
    request: StartSessionRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Start a new learning session."""
    session = Session(
        user_id=current_user.id,
        title=request.title,
        status="active"
    )
    db.add(session)
    await db.commit()
    await db.refresh(session)

    return StartSessionResponse(
        session_id=session.id,
        title=session.title,
        created_at=session.created_at
    )


@router.post("/{session_id}/message", response_model=SendMessageResponse)
async def send_message(
    session_id: UUID,
    request: SendMessageRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Send a message and receive AI response."""

    result = await db.execute(
        select(Session).where(
            Session.id == session_id,
            Session.user_id == current_user.id
        )
    )
    session = result.scalar_one_or_none()

    if not session:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Session not found"
        )

    history = await get_conversation_history(session_id, db)

    msg_count_result = await db.execute(
        select(func.count(Message.id)).where(Message.session_id == session_id)
    )
    msg_count = msg_count_result.scalar()

    await save_message(
        session_id=session_id,
        role="user",
        content=request.message,
        order=msg_count + 1,
        db=db
    )

    if session.status == "completed":
        # Roadmap already exists — use discussion mode
        roadmap_result = await db.execute(
            select(Roadmap).where(Roadmap.session_id == session_id)
        )
        existing_roadmap = roadmap_result.scalar_one_or_none()
        roadmap_context = {}
        if existing_roadmap:
            roadmap_context = {
                "goal": existing_roadmap.goal,
                "total_estimated_hours": existing_roadmap.total_estimated_hours,
                "weekly_hours": existing_roadmap.weekly_hours,
                "estimated_weeks": existing_roadmap.estimated_weeks,
                "skill_areas": existing_roadmap.skill_areas,
            }
        ai_response = await ai_service.process_followup_message(
            user_message=request.message,
            conversation_history=history,
            roadmap_data=roadmap_context
        )
    else:
        ai_response = await ai_service.process_message(
            user_message=request.message,
            conversation_history=history
        )

    ai_message_content = ai_response.get("message", "")

    await save_message(
        session_id=session_id,
        role="assistant",
        content=ai_message_content,
        order=msg_count + 2,
        db=db
    )

    roadmap_data = None
    if session.status != "completed" and ai_response.get("ready") and ai_response.get("roadmap"):
        roadmap_data = ai_response["roadmap"]
        await save_roadmap(session_id, roadmap_data, db)

        session.goal = roadmap_data.get("goal", "")
        session.status = "completed"
        session.title = roadmap_data.get("goal", session.title)

    await db.commit()

    return SendMessageResponse(
        session_id=session_id,
        ready=ai_response.get("ready", False),
        message=ai_message_content,
        roadmap=roadmap_data
    )


@router.get("/{session_id}/history", response_model=SessionHistoryResponse)
async def get_session_history(
    session_id: UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Get full conversation history for a session."""

    result = await db.execute(
        select(Session).where(
            Session.id == session_id,
            Session.user_id == current_user.id
        )
    )
    session = result.scalar_one_or_none()

    if not session:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Session not found"
        )

    msg_result = await db.execute(
        select(Message)
        .where(Message.session_id == session_id)
        .order_by(Message.order)
    )
    messages = msg_result.scalars().all()

    return SessionHistoryResponse(
        session_id=session.id,
        title=session.title,
        goal=session.goal,
        status=session.status,
        created_at=session.created_at,
        messages=[
            MessageResponse(
                role=msg.role,
                content=msg.content,
                created_at=msg.created_at
            )
            for msg in messages
        ]
    )


@router.delete("/{session_id}")
async def delete_session(
    session_id: UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Delete a session and all associated messages and roadmap."""

    result = await db.execute(
        select(Session).where(
            Session.id == session_id,
            Session.user_id == current_user.id
        )
    )
    session = result.scalar_one_or_none()

    if not session:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Session not found"
        )

    await db.execute(delete(Message).where(Message.session_id == session_id))
    await db.execute(delete(Roadmap).where(Roadmap.session_id == session_id))
    await db.delete(session)
    await db.commit()

    return {"message": "Session deleted successfully"}


@router.post("/upload-file")
async def upload_file(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """Extract text from an uploaded file (PDF, Word, Excel, or text)."""
    MAX_SIZE = 10 * 1024 * 1024  # 10 MB

    content = await file.read()
    if len(content) > MAX_SIZE:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File too large. Maximum size is 10 MB."
        )

    filename = file.filename or ""
    name_lower = filename.lower()
    extracted_text = ""

    try:
        if name_lower.endswith(".pdf"):
            if _PdfReader is None:
                raise HTTPException(status_code=status.HTTP_501_NOT_IMPLEMENTED, detail="PDF support not installed on server.")
            reader = _PdfReader(io.BytesIO(content))
            pages = [page.extract_text() or "" for page in reader.pages]
            extracted_text = "\n\n".join(p for p in pages if p.strip())

        elif name_lower.endswith(".docx"):
            if _DocxDocument is None:
                raise HTTPException(status_code=status.HTTP_501_NOT_IMPLEMENTED, detail="Word document support not installed on server.")
            doc = _DocxDocument(io.BytesIO(content))
            extracted_text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())

        elif name_lower.endswith((".xlsx", ".xls")):
            if _load_workbook is None:
                raise HTTPException(status_code=status.HTTP_501_NOT_IMPLEMENTED, detail="Excel support not installed on server.")
            wb = _load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            rows = []
            for sheet in wb.worksheets:
                rows.append(f"[Sheet: {sheet.title}]")
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join(str(c) if c is not None else "" for c in row)
                    if row_text.strip():
                        rows.append(row_text)
            extracted_text = "\n".join(rows)

        elif name_lower.endswith((".txt", ".md", ".csv", ".json")):
            try:
                extracted_text = content.decode("utf-8")
            except UnicodeDecodeError:
                extracted_text = content.decode("latin-1")

        else:
            raise HTTPException(
                status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
                detail="Unsupported file type. Supported: PDF, Word (.docx), Excel (.xlsx/.xls), text (.txt, .md, .csv, .json)"
            )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Could not extract text from file: {str(e)}"
        )

    if not extracted_text.strip():
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="No readable text could be extracted from this file."
        )

    MAX_CHARS = 50_000
    truncated = False
    if len(extracted_text) > MAX_CHARS:
        extracted_text = extracted_text[:MAX_CHARS]
        truncated = True

    return {
        "filename": filename,
        "content": extracted_text,
        "size": len(content),
        "truncated": truncated,
    }
